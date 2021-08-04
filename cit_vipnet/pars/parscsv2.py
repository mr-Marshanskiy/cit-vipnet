
from threading import Event
from openpyxl import load_workbook
from datetime import datetime, date
from events.models import Organisation, License, Vpn, Distributor, Network, Device
work_sheet = 'Лист1'
file_path = r'E:\db_new.xlsx'

class ParsXlsx():
    def __init__(self):
        self.wb = load_workbook(filename = file_path)
        self.ws = self.wb[work_sheet]    

    def unmerge_with_filling(self):
        mergedcells =[] #array of merged cells
        for group in self.ws.merged_cells.ranges:
            mergedcells.append(group)
        for group in mergedcells:
            # define bounds of merged cell
            min_col, min_row, max_col, max_row = group.bounds
            top_cell_value = self.ws.cell(row=min_row, column=min_col).value
            self.ws.unmerge_cells(str(group))
            for irow in range(min_row, max_row+1):
                self.ws.cell(row = irow, column = max_col, value = top_cell_value)
        return

    def save_new_xlsx(self):
        self.wb.save(r'E:\db_new.xlsx')

    def add_organisations(self):
        logs = []
        for row in self.ws.iter_rows(values_only=True):

            if row[6] is None:
                continue
            if not Organisation.objects.filter(
                inn=row[6],
            ).exists():
                try:
                    Organisation.objects.create(
                        reg_number=int(row[0]),
                        reg_date=row[1],
                        inn=row[6],
                        full_name=row[3],
                        short_name=row[3],
                        address=f'{row[4]} {row[7]} {row[8]}',
                        phone=row[9],
                        employee=row[10],
                    )
                except Exception as e:
                    logs.append(f'Организация {row[6]}  не была добавлена. Ошибка: {e}')
        return logs

    def add_licenses(self):
        logs = []
        for row in self.ws.iter_rows(values_only=True):
            if row[17] is None or row[17]=='':
                continue
            comment = ''
            if isinstance(row[18], (date, datetime)):
                lic_date = row[18]
            else:
                lic_date = date(2000, 1, 1)
                comment += f'Дата акта: {row[18]} | '
            try:
                ammount_convert = int(
                str(row[19]).strip().split(' ')[0]
            )
            except:
                ammount_convert = 777
                comment = f'Кол-во: {row[19]} |'
            
            if not row[16]:
                distr = 'Неизвестно'
            else:
                distr = row[16]

            if not License.objects.filter(
                act=row[17],
                distributor__name = distr,
            ).exists():
                try:
                    License.objects.create(
                        act=row[17],
                        date=lic_date,
                        distributor=Distributor.objects.get(
                                    name=row[16]),
                        amount=ammount_convert,
                        comment=comment,
                    )
                except Exception as e:
                    logs.append(f'Акт {row[17]}  не был добавлен. Ошибка: {e}')
        return logs

    def add_distributors(self):
        logs = []
        for row in self.ws.iter_rows(values_only=True):

            if row[16] is None:
                continue
            if not Distributor.objects.filter(
                name=row[16]
            ).exists():
                try:
                    Distributor.objects.create(
                        name=row[16],
                    )
                except Exception as e:
                    logs.append(f'Дистрибьютор {row[16]}  не был добавлен. Ошибка: {e}')
        return logs

    def add_devices(self):
        logs = []
        for row in self.ws.iter_rows(values_only=True):

            if row[14] is None:
                continue
            if not Device.objects.filter(
                type=row[14]
            ).exists():
                try:
                    Device.objects.create(
                        type=row[14],
                    )
                except Exception as e:
                    logs.append(f'Устройство {row[14]}  не было добавлено. Ошибка: {e}')
        return logs

    def add_vpn(self):
        logs = []
        for row in self.ws.iter_rows(values_only=True):
            if row[11] is None or not Organisation.objects.filter(inn=row[6]).exists():
                continue
            try:
                vpn = int(str(row[11]).split('-')[-1])
            except:
                vpn = 777

            org = Organisation.objects.prefetch_related().get(inn=row[6])
            if not org.orgs.filter(vpn_number=vpn, reg_number=row[12], network=row[2]).exists():
                try:
                    amount = int(str(row[19]).strip().split(' ')[0])
                except:
                    amount = 777
                lic = License.objects.filter(
                    act=row[17], distributor__name=row[16]).first()
                try:
                    Vpn.objects.create(
                        network=Network.objects.get(number=row[2]),
                        reg_number=row[12],
                        reg_date=row[13],
                        organisation=org,
                        vpn_number=vpn,
                        device_type=Device.objects.get(type=row[14]),
                        device_id=row[15],
                        license=lic,
                    )
                except Exception as e:
                    logs.append(e)
                    try:
                        Vpn.objects.create(
                            reg_number=row[12],
                            reg_date=row[13],
                            organisation=org,
                            vpn_number=vpn,
                            device_id=row[15],
                            comment = (f'Сеть: {row[2]}\n'
                                        f'Устройство: {row[14]}\n'
                                        f'Дистр: {row[16]}\n'
                                        f'Номер акта: {row[17]}\n'
                                        f'Дата акта: {row[18]}\n'
                                        f'Кол-во: {row[19]}\n') 
                        )
                    except Exception as e:
                        logs.append(f'СКИ {row[11]} ({row[6]}) не была добавлена. Ошибка: {e}')
        return logs
